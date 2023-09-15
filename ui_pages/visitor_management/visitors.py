__author__ = 'prahlad.sharma'

import csv
import xlrd
from pandas.tests.io.excel.test_xlsxwriter import xlsxwriter
from selenium.common.exceptions import ElementClickInterceptedException

from virgin_utils import *
import openpyxl
import pathlib
import pandas as pd


def import_visitors():
    """
    Import the visitors
    """
    visitors_file_path = os.getcwd() + '/test_data/' + 'visitor.xlsx'
    df = pd.DataFrame({'First Name': ['Jack'], 'Last Name': ['martin'], 'Citizenship': ['India'], 'Gender': ['Female'],
                       'Date Of Birth': ['01/01/1975'], 'Contact No.': ['343433434'], 'Company Name': ['DXP'],
                       'Employee Number': ['98'], 'Department': ['Deck'],
                       'Visitor Notes': ['Uploaded via bulk upload'],
                       'Visitor Type': ['Vendor or Contractor'], 'Reason of Visit': ['Decor Install'],
                       'Boarding Type': ['General Boarding'],
                       'Visit Start Date': ['12/08/2020'], 'License Plate Number': [''], 'ID Type': ['Passport'],
                       'ID Number': ['222222222'],
                       'Shipboard Contact': ['john']})
    df.to_excel(visitors_file_path, index=False)

    writer = pd.ExcelWriter(visitors_file_path, engine='xlsxwriter')
    df.to_excel(writer, sheet_name='Sheet1', index=False)
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    format1 = workbook.add_format({'num_format': 'mm/dd/yy', 'align': 'left'})
    # format2 = workbook.add_format({'num_format': '0%'})
    worksheet.set_column('B:B', 18, format1)
    worksheet.set_column('N:N', 18, format1)
    writer.save()

    # workbook = openpyxl.load_workbook(visitors_file_path)
    # sheet = workbook.active
    # sheet['A2'] = first_name
    # sheet['B2'] = last_name
    # sheet['C2'] = country
    # sheet['Q2'] = doc_id
    # sheet['N2'] = visit_date
    # workbook.template = False
    # workbook.save(visitors_file_path)
    # workbook.close()
    # self.webDriver.set_text(element=self.locators.import_directly, locator_type='xpath', text=visitors_file_path)
    # self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
    # self.webDriver.enter_data_in_textbox_using_action(element=self.locators.reviewer,
    #                                                   locator_type='xpath',
    #                                                   text='Vertical QA')
    # self.webDriver.click(element=self.locators.import_in_popup, locator_type='xpath')
    # self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
    # self.verify_visitor_get_added(first_name, last_name, country)


class Visitors(General):
    """
    Page Class for Visitors page
    """

    def __init__(self, web_driver):
        """
        To Initialize the locators of Visitor page
        :param web_driver:
        """
        super().__init__()
        self.visitor_data = dict()
        self.webDriver = web_driver
        self.locators = self.dict_to_ns({
            "loader": "//div[@class='LoaderContainer__div']//img",
            "page_header_1": "//span[contains(text(),'%s')]",
            "page_header": "//p[@class='pageTitle']//span[contains(text(),'%s')]",
            "add_visitor": "//span[contains(text(),'NEW VISITOR')]",
            "import": "//span[contains(text(),'IMPORT')]",
            "import_directly": "//span[contains(text(),'IMPORT')]/following-sibling::input",
            "export_icon": "//div[@id='DownloadContainer__div']//img[contains(@class,'SVGContainer')]",
            "export_excel": "//span[contains(text(),'Export Excel')]",
            "export_csv": "//span[contains(text(),'Export CSV')]",
            "filter": "//div[@class='filters visitorFilter']//img[contains(@class,'SVGContainer')]",
            "action_dropdown": "//div[@id='ActionsDropdown']",
            "firstname": "//input[@name='firstName']",
            "click_visits_section": "//button[text()='VISITS']",
            "lastname": "//input[@name='lastName']",
            "DOB": "//div[@class='react-datepicker__input-container']//input",
            "gender": "//div[@id='genderCode']//div[@class='css-1g6gooi']",
            "phone": "//input[@name='contactNumber']",
            "citizenship": "//div[@id='citizenShipCountryCode']//div[@class='css-1g6gooi']",
            "company": "//div/input[@name='companyName']",
            "id_type": "//div[@id='documentTypeCode']//div[@class='css-1g6gooi']",
            "id_number": "//input[@name='number']",
            "emp_number": "//input[@name='employeeNumber']",
            "notes": "//textarea[@name='visitorNotes']",
            "visitor_request_type": "//div[@id='visitorTypeCode']//div[@class='css-1g6gooi']",
            "departments": "//div[@id='departmentId']//div[@class='react-select__value-container css-1hwfws3']",
            "department_poc": "//div[contains(@class,'textfield')]//input[@name='departmentPOC']",
            "add_visit": "//span[contains(text(),'ADD VISIT')]",
            "visit_date": "//div[@class='column']//div[@class='react-datepicker__input-container']/input",
            "visit_purpose": "//textarea[@name='purpose']",
            "boarding_type": "//div[@id='boardingTypeCode']//div[@class='css-1g6gooi']",
            "reviewer": "//div[@id='reviewerTeamMemberId']//div[@class='css-1g6gooi']",
            "cancel": "//button[@class='button is-default is-fullwidth ']/span[contains(text(),'CANCEL')]",
            "send_for_approval": "//span[contains(text(),'SEND FOR APPROVAL')]",
            "submit": "//button[@class='button is-primary is-fullwidth ']",
            "photo_path": "//div[@class='column photo is-one-quarter false']//input",
            "import_in_popup": "//button[contains(@class,'button is-primary')]//span[contains(text(),'IMPORT')]",
            "visitor_checkbox": "//body[@class='vv-light']/div[@id='root']/div[@class='RootContainer']/div[@class='LayoutContainer']/div[@class='PageContainer']/div[@class='section']/div[@class='VisitorsContainer__div container']/div[@class='card']/div[contains(@class,'table-container')]/table[@class='table is-hoverable is-fullwidth visitorList-table']/thead/tr/th/span[contains(@class,'checkbox')]/span[@class='tick']/img",
            "action_dd": "//span[contains(text(),'Actions')]",
            "approve_visitor": "//a[@name='APPROVE']",
            "port_name": "//input[@name='portCode']",
            "visitor_name_in_table": "//div[@class='ProfileContainer is-pulled-left']/img/..",
            "visitor_visit_date_in_table": "//tr[%s]//td[%s]",
            "visitor_citizenship_in_table": "//tr[%s]//td[%s]",
            "view_visitor_details_arrow": "//tr[%s]//td[%s]//a[1]//button[1]//img[1]",

            "visitor_details": "//span[contains(text(),'Visitor Details')]",
            "details_visitor_name": "//div[@class='column is-full title']",
            "details_visit_date": "",
            "details_visitor_citizonship": "//span[contains(text(),'Citizenship')]/../following-sibling::div",
            "details_gender": "//span[contains(text(),'Gender')]/../following-sibling::div",
            "details_ID_type": "//span[contains(text(),'ID Type')]/../following-sibling::div",
            "details_ID_Number": "//span[contains(text(),'ID Number')]/../following-sibling::div",
            "details_department": "//span[contains(text(),'Department')]/../following-sibling::div",
            "details_employeeNo": "//span[contains(text(),'Employee No')]/../following-sibling::div",
            "details_visitor_type": "//span[contains(text(),'Visitor/Request Type')]/../following-sibling::div",
            "details_notes": "//span[contains(text(),'Notes for Visitor')]/../following-sibling::div",

            "edit_visitor": "//button[contains(@class,'button column is-icon-button has-text-right')]//img",
            "save_visitor": "//span[contains(text(),'SAVE')]",

            "visits_tab": "//button[contains(text(),'VISITS')]",
            "alert_tab": "//button[contains(text(),'ALERTS')]",
            "message_tab": "//button[contains(text(),'MESSAGES')]",

            "visit_status_in_table": "//div[@class='status']",
            "visit_arrow": "//tbody//img[contains(@class,'SVGContainer')]",
            "click_visit_arrow": "//tbody//button/img[contains(@class,'SVGContainer')]",
            "visit_status": "//div[@class='text status']",
            "approve_visit": "//span[contains(text(),'APPROVE')]",
            "reject_visit": "//span[contains(text(),'REJECT')]",
            "rejection_reason": "//span[contains(text(),'Invalid Details')]",
            "details_3_dots": "//div[@id='MoreOption__div']//img[contains(@class,'SVGContainer')]",
            "more_option_3_dot": "//div[@id='visitorList']//button[contains(@class,'button is-icon-button "
                                 "dropdown-trigger')]//img[contains(@class,'SVGContainer')]",
            "create_alert": "//span[contains(text(),'Create Alert')]",
            "create_message": "//span[contains(text(),'Create Message')]",
            "remove_visitor": "//span[contains(text(),'Remove Visitor')]",
            "alert_type": "//span[contains(text(),'Alert Type')]",
            "alert_department": "//span[contains(text(),'Department')]",
            "alert_details": "//textarea[@name='description']",
            "alert_create": "//span[contains(text(),'CREATE')]",
            "dynamic_xpath": "//td[text()='%s']",
            "get_message_description": "//span[text()='Description']/ancestor::table/tbody/tr/td[contains(text(),'message for')]",
            "get_alert_description": "//span[text()='Description']/ancestor::table/tbody/tr/td[contains(text(),'alert for')]",
            "visit_purpose_value": "//div[@class='label']//span[contains(text(),'Visit "
                                   "Purpose')]/../following-sibling::div",
            "arrow": "//div[@class='TabContent data active']//tbody//button[1]//img[1]",
            "delete": "//div[@class='TabContent data active']//button[2]//img[1]",
            "edit_alert_message": "//span[contains(text(),'EDIT')]",
            "delete_alert_message": "//span[contains(text(),'DELETE')]",
            "update_alert_message": "//span[contains(text(),'UPDATE')]",
            "confirmation_delete": "//button[contains(@class,'button is-primary')]//span[contains(text(),'Delete')]",
            "confirmation_cancel": "//span[contains(text(),'Cancel')]",
            "alert_message_description": "//span[contains(text(),'Description')]/../following-sibling::div",
            "rows": "//tr[%s]",
            "table_row": "//tr",
            "table_rows_without_header": "//tbody/tr",
            "no_records": "//span[contains(text(),'No Record(s) Available.')]",
            "new_visitor_disabled_button": "//button[@disabled and @class='button is-default ']/span[contains(text(),"
                                           "'NEW VISITOR')] ",
            "column_name_in_3_dots": "//label[contains(text(),'%s')]",
            "header_name": "//thead/tr/th",
            "column_value_in_row": "//tbody/tr[%s]/td[%d]",
            "success_toast": "//div[@class='notification is-success']",
            "visitors": "//span[@class='checkbox is-pulled-left ']",
            "visitors_last_name": "//div[@class='ProfileContainer is-pulled-left']/../following-sibling::td[%s]",
            "apply": "//span[contains(text(),'APPLY')]",
        })

    def verify_toast_title(self):
        """
        Function to verify the success toast
        """
        screen_title = self.webDriver.get_text(element=self.locators.success_toast, locator_type='xpath')
        if screen_title == "Visitor is successfully APPROVED.":
            self.webDriver.allure_attach_jpeg('valid_toast')
            logger.debug(f"correct message is display and message is {screen_title}")
        else:
            self.webDriver.allure_attach_jpeg('error_invalid_toast')
            raise Exception(f"correct message is not display and message is {screen_title}")

    def verification_of_page_header(self, page_header):
        """
        To check the availability of Visitor page
        :param page_header:
        """
        if page_header in ('Visitors', 'Create Alert', 'Create Message', 'Change Boarding Slot'):
            self.webDriver.explicit_visibility_of_element(element=self.locators.page_header_1 % page_header,locator_type='xpath',time_out=60)
            screen_title = self.webDriver.get_text(element=self.locators.page_header_1 % page_header,locator_type='xpath')
            if screen_title == page_header:
                logger.debug(f"User is landed on {page_header} page")
                self.webDriver.allure_attach_jpeg(f"{page_header}")
            else:
                self.webDriver.allure_attach_jpeg(f"{page_header}_error")
                raise Exception(f"User is not landed on {page_header} page")
        else:
            self.webDriver.wait_till_element_appear_on_screen(element=self.locators.page_header % page_header,
                                                              locator_type='xpath')
            screen_title = self.webDriver.get_text(element=self.locators.page_header % page_header,
                                                   locator_type='xpath')
            if screen_title == page_header:
                logger.debug(f"User is landed on {page_header} page")
                self.webDriver.allure_attach_jpeg("visitor")
            else:
                self.webDriver.allure_attach_jpeg("error_visitor")
                raise Exception(f"User is not landed on {page_header} page")

    def verify_export(self, file_name):
        """
        Function to click on export functionality
        :param file_name:
        """
        self.webDriver.allure_attach_jpeg('before_export_click')
        self.webDriver.scroll_complete_page_top()
        self.webDriver.wait_for(1)
        self.webDriver.click(element=self.locators.export_icon, locator_type='xpath')
        self.webDriver.wait_for(1)
        self.webDriver.click(element=self.locators.export_csv, locator_type='xpath')
        self.webDriver.wait_for(1)
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('after_export_click')

        if os.environ.get('EXECUTOR_NUMBER', None):
            flag = self.webDriver.grid_download_verification(file_name=file_name)
        else:
            download_path = os.getcwd() + '/downloaded_files/'
            if not os.path.isdir(download_path):
                os.makedirs(download_path)
            file_path = pathlib.Path(download_path + f"{file_name}.csv")
            flag = self.webDriver.check_local_system_download(file_path)

        if flag:
            logger.debug("file is downloaded successfully")
        else:
            raise Exception("File is not downloaded successfully")

    def click_filter_option(self):
        """
        Function to click on filter option
        """
        self.webDriver.click(element=self.locators.filter, locator_type='xpath')
        self.webDriver.wait_for(2)
        self.webDriver.allure_attach_jpeg("filer_open_visitor")

    def click_add_visitor(self):
        """
        Function to click on add visitor button
        """
        self.webDriver.wait_for(2)
        self.webDriver.click(element=self.locators.add_visitor, locator_type='xpath')

    def fill_visitor_form_and_save(self, visitor_first_name, visitor_last_name, citizenship, reviewer, ship_visit_date,
                                   test_data, side):
        """
        Function to fill the visitor form
        :param visitor_first_name:
        :param visitor_last_name:
        :param citizenship:
        :param reviewer:
        :param ship_visit_date:
        :param test_data:
        """
        self.visitor_data["visitorFirstName"] = visitor_first_name
        self.visitor_data["visitorLastName"] = visitor_last_name
        self.visitor_data["citizenship"] = citizenship
        visitor_image_path = GeneratePhoto(gender='M').select_random_image()
        self.webDriver.set_text(element=self.locators.photo_path, locator_type='xpath', text=visitor_image_path)
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.set_text(element=self.locators.firstname, locator_type='xpath', text=visitor_first_name)
        self.webDriver.set_text(element=self.locators.lastname, locator_type='xpath', text=visitor_last_name)
        self.webDriver.set_text(element=self.locators.DOB, locator_type='xpath', text="09/08/1999")
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.gender,
                                                          locator_type='xpath',
                                                          text='Female')
        self.webDriver.set_text(element=self.locators.phone, locator_type='xpath', text=generate_phone_number(10))
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.citizenship,
                                                          locator_type='xpath',
                                                          text=citizenship)
        self.webDriver.set_text(element=self.locators.company, locator_type='xpath', text="Decurtis")
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.id_type,
                                                          locator_type='xpath',
                                                          text='Passport')
        self.visitor_data["id_number"] = generate_phone_number(8)
        self.webDriver.set_text(element=self.locators.id_number, locator_type='xpath',
                                text=self.visitor_data["id_number"])
        self.webDriver.set_text(element=self.locators.emp_number, locator_type='xpath', text=generate_phone_number(8))
        self.webDriver.set_text(element=self.locators.notes, locator_type='xpath', text=generate_phone_number(8))
        self.webDriver.scroll_till_element(element=self.locators.visitor_request_type, locator_type='xpath')
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.visitor_request_type,
                                                          locator_type='xpath',
                                                          text='Employee Business')
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.departments,
                                                          locator_type='xpath',
                                                          text='Engine')
        self.webDriver.allure_attach_jpeg("add_visitor_1")
        self.webDriver.wait_for(10)
        self.webDriver.set_text(element=self.locators.department_poc, locator_type='xpath', text='test')
        self.webDriver.allure_attach_jpeg("add_visitor_2")
        self.webDriver.click(element=self.locators.add_visit, locator_type='xpath')
        self.webDriver.scroll_till_element(element=self.locators.visit_date, locator_type='xpath')
        self.webDriver.wait_for(2)
        self.webDriver.set_text(element=self.locators.visit_date, locator_type='xpath', text=ship_visit_date)
        self.webDriver.set_text(element=self.locators.visit_purpose, locator_type='xpath', text='testing')
        test_data['port_name'] = self.webDriver.get_text(element=self.locators.port_name, locator_type='xpath')
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.boarding_type,
                                                          locator_type='xpath',
                                                          text='General Boarding')
        if side == 'shore':
            self.webDriver.enter_data_in_textbox_using_action(element=self.locators.reviewer,
                                                              locator_type='xpath',
                                                              text=reviewer)
        self.webDriver.allure_attach_jpeg("add_visitor_3")

        self.webDriver.click(element=self.locators.submit, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg("loader_disappear")

    def get_count_on_list(self):
        """
        Function to get the count on list
        """
        flag = False
        if self.webDriver.is_element_display_on_screen(element=self.locators.no_records, locator_type='xpath'):
            return 0
        else:
            while not flag:
                row_count = len(
                    self.webDriver.get_elements(element=self.locators.table_rows_without_header, locator_type='xpath'))
                try:
                    self.webDriver.explicit_click(element=self.locators.rows % (row_count - 1), locator_type='xpath')
                    self.webDriver.wait_for(3)
                except ElementClickInterceptedException:
                    try:
                        self.webDriver.wait_for(5)
                        self.webDriver.explicit_click(element=self.locators.rows % (row_count - 1),
                                                      locator_type='xpath')
                    except ElementClickInterceptedException:
                        self.webDriver.wait_for(2)
                        self.webDriver.explicit_click(element=self.locators.rows % (row_count - 1),
                                                      locator_type='xpath')

                self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader,
                                                                       locator_type='xpath')
                updated_rows = len(
                    self.webDriver.get_elements(element=self.locators.table_rows_without_header, locator_type='xpath'))
                if row_count < updated_rows:
                    flag = False
                else:
                    flag = True
            return len(
                self.webDriver.get_elements(element=self.locators.table_rows_without_header, locator_type='xpath'))

    def add_new_column_in_list(self, column_name):
        """
        Function to add the new column
        :param column_name:
        """
        self.open_more_option_drawer()
        self.webDriver.explicit_click(element=self.locators.column_name_in_3_dots % column_name, locator_type='xpath')
        self.webDriver.click(element=self.locators.apply, locator_type='xpath')

    def set_column_name(self, test_data):
        """
        Function to get the column name
        :param test_data:
        """
        column_number = 0
        total_column = self.webDriver.get_elements(element=self.locators.header_name,
                                                   locator_type='xpath')
        for column in total_column:
            column_number = column_number + 1
            test_data[f"{column.text}_position"] = column_number

    def verify_visitor_get_added(self, visitor_first_name, visitor_last_name, citizenship, diff, test_data,
                                 ship_visit_date):
        """
        To verify the visitor after save
        :param visitor_first_name:
        :param visitor_last_name:
        :param citizenship:
        :param ship_visit_date:
        :param diff:
        :param test_data:
        """
        count = 0
        column_number = 0
        flag = False
        while not flag:
            row_count = len(self.webDriver.get_elements(element=self.locators.table_row, locator_type='xpath'))
            try:
                self.webDriver.explicit_click(element=self.locators.rows % (row_count - 1), locator_type='xpath')
                self.webDriver.wait_for(5)
            except ElementClickInterceptedException:
                self.webDriver.wait_for(10)
                self.webDriver.explicit_click(element=self.locators.rows % (row_count - 1), locator_type='xpath')

            self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
            self.webDriver.wait_for(5)
            updated_rows = len(self.webDriver.get_elements(element=self.locators.table_row, locator_type='xpath'))
            if row_count < updated_rows:
                flag = False
            else:
                flag = True
        visitor_name_list = self.webDriver.get_elements(element=self.locators.visitor_name_in_table,
                                                        locator_type='xpath')
        total_column = self.webDriver.get_elements(element=self.locators.header_name,
                                                   locator_type='xpath')
        for column in total_column:
            column_number = column_number + 1
            test_data[f"{column.text}_position"] = column_number

        if diff == 'NAME':
            for name in visitor_name_list:
                count = count + 1
                if name.text == visitor_first_name + " " + visitor_last_name:
                    logger.debug("Name is matching")
                    self.webDriver.scroll_till_element(
                        element=self.locators.view_visitor_details_arrow % (count, test_data['_position']),
                        locator_type='xpath')
                    self.webDriver.allure_attach_jpeg('visitor_name')
                    break
            else:
                self.webDriver.allure_attach_jpeg('error_visitor_name')
                raise Exception(f"Visitor name is not matching {visitor_first_name} {visitor_last_name}")

        else:
            for id in visitor_name_list:
                count = count + 1
                id_value = self.webDriver.get_text(
                    element=self.locators.column_value_in_row % (count, test_data['ID Number_position']),
                    locator_type='xpath')
                if str(id_value) == str(self.visitor_data["id_number"]):
                    self.webDriver.scroll_till_element(
                        element=self.locators.view_visitor_details_arrow % (count, test_data['_position']),
                        locator_type='xpath')
                    self.webDriver.allure_attach_jpeg('visitor_name')
                    break
            else:
                self.webDriver.allure_attach_jpeg('error_visitor_name')
                raise Exception(f"Visitor name is not matching {visitor_first_name} {visitor_last_name}")

        visit_date = self.webDriver.get_text(
            element=self.locators.visitor_visit_date_in_table % (count, test_data['Visit Start Date_position']),
            locator_type='xpath')
        visitor_citizenship = self.webDriver.get_text(
            element=self.locators.visitor_citizenship_in_table % (count, test_data['Citizenship_position']),
            locator_type='xpath')
        assert visit_date == ship_visit_date, 'Visit date is not matching'
        assert visitor_citizenship == citizenship, 'Visitor citizenship is not matching'
        self.visitor_data["visitorPosition"] = count

    def view_visitor(self, test_data):
        """
        Function to view the visitor details
        """
        self.webDriver.click(element=self.locators.view_visitor_details_arrow % (self.visitor_data["visitorPosition"],
                                                                                 test_data['_position']),
                             locator_type='xpath')
        self.webDriver.wait_till_element_appear_on_screen(element=self.locators.visitor_details,
                                                          locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')

    def view_first_visitor(self, test_data):
        """
        Function to view the visitor details
        """
        self.webDriver.click(element=self.locators.view_visitor_details_arrow % (1, test_data['_position']),
                             locator_type='xpath')
        self.webDriver.wait_till_element_appear_on_screen(element=self.locators.visitor_details,
                                                          locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')

    def verify_visitor_details_on_details_screen(self, visitor_first_name, visitor_last_name, citizenship):
        """
        Function to verify the visitor details on visitor details screen
        :param visitor_first_name:
        :param visitor_last_name:
        :param citizenship:
        """
        visitor_name = self.webDriver.get_text(element=self.locators.details_visitor_name, locator_type='xpath')
        if visitor_name == visitor_first_name + " " + visitor_last_name:
            self.webDriver.allure_attach_jpeg('visitor_name_details')
            logger.debug("Visitor name is matching on visitor details page")
        else:
            self.webDriver.allure_attach_jpeg('error_visitor_name_details')
            raise Exception("Visitor name is not matching on visitor details page")

        visitor_citizenship = self.webDriver.get_text(element=self.locators.details_visitor_citizonship,
                                                      locator_type='xpath')
        if visitor_citizenship == citizenship:
            logger.debug("Citizenship is matching on visitor details page")
        else:
            raise Exception("Citizenship is not matching on visitor details page")

        id_number = self.webDriver.get_text(element=self.locators.details_ID_Number, locator_type='xpath')
        if int(id_number) == int(self.visitor_data["id_number"]):
            logger.debug("ID number is matching on visitor details page")
        else:
            raise Exception("ID number is not matching on visitor details page")

    def click_edit_visitor(self):
        """
        Function to edit the Visitor
        """
        self.webDriver.click(element=self.locators.edit_visitor, locator_type='xpath')

    def select_all_visitor(self):
        """
        Function to select all visitor
        """
        self.webDriver.click(element=self.locators.visitor_checkbox, locator_type='xpath')

    def select_custom_visitors(self, last_name_position, limit):
        """
        Function to select custom visitor
        :param last_name_position
        :param limit
        """
        last_name = []
        visitors = self.webDriver.get_elements(element=self.locators.visitors, locator_type='xpath')
        if len(visitors) < limit:
            limit = len(visitors)
        visitors_first_name = self.webDriver.get_elements(
            element=self.locators.visitors_last_name % last_name_position,
            locator_type='xpath')
        for _visitor in visitors:
            _visitor.click()
            self.webDriver.wait_for(1)
            last_name.append(visitors_first_name[visitors.index(_visitor)].text)
            if visitors.index(_visitor) >= limit:
                break
        if len(visitors) > limit:
            self.select_all_visitor()

        return last_name

    def select_custom_visitors_by_id_number(self, id_number_position, limit, test_data):
        """
        Function to select custom visitor
        :param id_number_position
        :param limit
        """
        id_numbers = []
        visitors = self.webDriver.get_elements(element=self.locators.visitors, locator_type='xpath')
        if len(visitors) < limit:
            limit = len(visitors)
        visitors_id_number = self.webDriver.get_elements(
            element=self.locators.visitors_last_name % id_number_position,
            locator_type='xpath')
        count = 1
        for _visitor in visitors:
            _visitor.click()
            self.webDriver.wait_for(1)
            id_numbers.append(
                self.webDriver.get_text(element=self.locators.column_value_in_row % (count, id_number_position),
                                        locator_type='xpath'))
            count = count + 1
            if visitors.index(_visitor) >= limit:
                break
        return id_numbers

    def all_approve(self):
        """
        Approve the all visitor
        """
        self.webDriver.click(element=self.locators.action_dd, locator_type='xpath')
        self.webDriver.wait_for(1)
        self.webDriver.click(element=self.locators.approve_visitor, locator_type='xpath')

    def edit_visitor(self):
        """
        Function to edit the Visitor
        """
        self.webDriver.allure_attach_jpeg('edit_visitor')
        edited_visitor_first_name = generate_first_name()
        self.visitor_data["updatedVisitorName"] = edited_visitor_first_name
        self.webDriver.clear_text(element=self.locators.firstname, locator_type='xpath', action_type='clear')
        self.webDriver.set_text(element=self.locators.firstname, locator_type='xpath', text=edited_visitor_first_name)
        self.webDriver.click(element=self.locators.save_visitor, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        visitor_name = self.webDriver.get_text(element=self.locators.details_visitor_name, locator_type='xpath')
        if visitor_name == f"{self.visitor_data['updatedVisitorName']} {self.visitor_data['visitorLastName']}":
            self.webDriver.allure_attach_jpeg('updated_visitor_name')
            logger.debug("After update the first name, name is matching on visitor details page")
        else:
            self.webDriver.allure_attach_jpeg('error_updated_visitor_name')
            raise Exception(
                f"After update the first name, name is not matching on visitor details page, display Name: "
                f"{visitor_name} and expected Name : {self.visitor_data['updatedVisitorName']} "
                f"{self.visitor_data['visitorLastName']}")

        return edited_visitor_first_name

    def open_visit_tab(self):
        """
        Function to open the Visit tab
        """
        self.webDriver.click(element=self.locators.visits_tab, locator_type='xpath')
        self.webDriver.scroll_complete_page()

    def open_alert_tab(self):
        """
        Function to open the Alert tab
        """
        self.webDriver.click(element=self.locators.alert_tab, locator_type='xpath')
        self.webDriver.scroll_complete_page()

    def open_message_tab(self):
        """
        Function to open the Message tab
        """
        self.webDriver.click(element=self.locators.message_tab, locator_type='xpath')
        self.webDriver.scroll_complete_page()

    def click_cancel(self):
        """
        Function to click on cancel button
        """
        self.webDriver.click(element=self.locators.cancel, locator_type='xpath')

    def get_visit_status(self):
        """
        Function to check the status in popup
        """
        self.webDriver.click(element=self.locators.click_visit_arrow, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('status')
        return self.webDriver.get_text(element=self.locators.visit_status, locator_type='xpath')

    def get_visit_status_in_visitor_details(self):
        """
        Function to check the status in visit table
        """
        self.webDriver.allure_attach_jpeg('visitor_details')
        self.webDriver.scroll_complete_page_top()
        return self.webDriver.get_text(element=self.locators.visit_status_in_table, locator_type='xpath')

    def approve_visit(self):
        """
        Function to approve the Visit
        """
        self.webDriver.click(element=self.locators.approve_visit, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')

    def reject_visit(self):
        """
        Function to reject the Visit
        """
        self.webDriver.click(element=self.locators.click_visits_section, locator_type='xpath')
        self.webDriver.wait_for(3)
        self.webDriver.click(element=self.locators.click_visit_arrow, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('reject_reason1')
        self.webDriver.wait_for(3)
        self.webDriver.click(element=self.locators.reject_visit, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('reject_reason2')
        self.webDriver.wait_for(3)
        self.webDriver.click(element=self.locators.rejection_reason, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('reject_reason3')
        self.webDriver.click(element=self.locators.reject_visit, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')

    def open_more_option_drawer(self):
        """
        Function to open the more option/3 dot menu
        """
        self.webDriver.scroll_complete_page_top()
        self.webDriver.click(element=self.locators.more_option_3_dot, locator_type='xpath')

    def open_more_option_drawer_details(self):
        """
        Function to open the more option/3 dot menu
        """
        self.webDriver.scroll_complete_page_top()
        self.webDriver.click(element=self.locators.details_3_dots, locator_type='xpath')

    def click_create_alert(self):
        """
        Function to open the create alert button
        """
        self.webDriver.click(element=self.locators.create_alert, locator_type='xpath')
        self.webDriver.wait_till_element_appear_on_screen(element=self.locators.alert_type,
                                                          locator_type='xpath')

    def create_alert(self, alert_description):
        """
        Function to create the alert
        :param alert_description:
        """
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.alert_type,
                                                          locator_type='xpath',
                                                          text='Allow Onboard')
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.alert_department,
                                                          locator_type='xpath',
                                                          text='Engine')
        self.webDriver.set_text(element=self.locators.alert_details, locator_type='xpath', text=alert_description)
        self.webDriver.allure_attach_jpeg('create_alert')
        self.webDriver.click(element=self.locators.alert_create, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')

    def verification_of_alert_after_creation(self, alert_description):
        """
        Function to verify the alert after create
        :param alert_description:
        """
        self.open_alert_tab()
        if alert_description == self.webDriver.get_text(element=self.locators.get_alert_description,
                                                        locator_type='xpath'):
            self.webDriver.allure_attach_jpeg('alert_desc')
            logger.debug("Alert Description is matching after creation ")
        else:
            self.webDriver.allure_attach_jpeg('error_alert_desc')
            raise Exception("Alert Description is not matching after creation ")

    def verification_of_alert_on_details_screen(self, alert_description):
        """
        Function to verify the alert on details screen
        :param alert_description:
        """
        self.open_alert_tab()
        self.webDriver.click(element=self.locators.arrow, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.wait_till_element_appear_on_screen(element=self.locators.alert_message_description,
                                                          locator_type='xpath')
        if alert_description == self.webDriver.get_text(element=self.locators.alert_message_description,
                                                        locator_type='xpath'):
            self.webDriver.allure_attach_jpeg('alert_desc_on_details')
            logger.debug("Alert Description is matching after creation on details page")
        else:
            self.webDriver.allure_attach_jpeg('error_alert_desc_on_details')
            raise Exception("Alert Description is not matching after creation on details page")

    def verification_of_alert_description_after_creation(self, alert_description):
        """
        Function to verify the alert after create
        :param alert_description:
        """
        self.open_alert_tab()
        if alert_description == self.webDriver.get_text(element=self.locators.dynamic_xpath % alert_description,
                                                        locator_type='xpath'):
            self.webDriver.allure_attach_jpeg('alert_desc')
            logger.debug("Alert Description is matching after creation ")
        else:
            self.webDriver.allure_attach_jpeg('error_alert_desc')
            raise Exception("Alert Description is not matching after creation ")


    def edit_alert_and_verify(self, edit_alert_description):
        """
        Function to edit the alert
        :param edit_alert_description:
        """
        self.open_more_option_drawer_details()
        self.webDriver.allure_attach_jpeg('more_option')
        self.webDriver.click(element=self.locators.edit_alert_message, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.clear_text(element=self.locators.alert_details, locator_type='xpath', action_type='clear')
        self.webDriver.set_text(element=self.locators.alert_details, locator_type='xpath', text=edit_alert_description)
        self.webDriver.allure_attach_jpeg('edited_alert')
        self.webDriver.click(element=self.locators.update_alert_message, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('update_alert')
        self.webDriver.navigate_back()
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.open_alert_tab()
        self.webDriver.allure_attach_jpeg('alert_updated_table')
        self.verification_of_alert_after_creation(edit_alert_description)

    def delete_alert(self):
        """
        Function to delete the alert
        """
        self.webDriver.click(element=self.locators.delete, locator_type='xpath')
        self.webDriver.click(element=self.locators.confirmation_delete, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('delete_alert')

    def edit_message_and_verify(self, edit_message_description):
        """
        Function to edit the alert
        :param edit_message_description:
        """
        self.open_more_option_drawer_details()
        self.webDriver.allure_attach_jpeg('message_drawer')
        self.webDriver.click(element=self.locators.edit_alert_message, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.clear_text(element=self.locators.alert_details, locator_type='xpath', action_type='clear')
        self.webDriver.set_text(element=self.locators.alert_details, locator_type='xpath',
                                text=edit_message_description)
        self.webDriver.allure_attach_jpeg('fill_updated_message')
        self.webDriver.click(element=self.locators.update_alert_message, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('message_details')
        self.webDriver.navigate_back()
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.open_message_tab()
        self.webDriver.allure_attach_jpeg('edited_message')
        self.verification_of_message_after_creation(edit_message_description)

    def verification_of_message_description_after_creation(self, message_description):
        """
        Function to verify the alert after create
        :param message_description:
        """
        self.open_message_tab()
        if message_description == self.webDriver.get_text(element=self.locators.dynamic_xpath % message_description,
                                                          locator_type='xpath'):
            self.webDriver.allure_attach_jpeg('added_message')
            logger.debug("Message Description is matching after creation ")
        else:
            self.webDriver.allure_attach_jpeg('error_added_message')
            raise Exception("Message Description is not matching after creation ")

    def verification_of_message_on_details_screen(self, message_description):
        """
        Function to verify the alert on details screen
        :param message_description:
        """
        self.open_message_tab()
        self.webDriver.click(element=self.locators.arrow, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        if message_description == self.webDriver.get_text(element=self.locators.alert_message_description,
                                                          locator_type='xpath'):
            self.webDriver.allure_attach_jpeg('message_desc_on_details')
            logger.debug("Message Description is matching after creation on details page")
        else:
            self.webDriver.allure_attach_jpeg('error_message_desc_on_details')
            raise Exception("Message Description is not matching after creation on details page")

    def delete_message(self):
        """
        Function to delete the alert
        """
        self.webDriver.click(element=self.locators.delete, locator_type='xpath')
        self.webDriver.click(element=self.locators.confirmation_delete, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('deleted_message')

    def click_create_message(self):
        """
        Function to open the create message button
        """
        self.webDriver.click(element=self.locators.create_message, locator_type='xpath')

    def create_message(self, message_description):
        """
        Function to create the message
        :param message_description:
        """
        self.webDriver.set_text(element=self.locators.alert_details, locator_type='xpath', text=message_description)
        self.webDriver.allure_attach_jpeg('fill_message')
        self.webDriver.click(element=self.locators.alert_create, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')

    def verification_of_message_after_creation(self, message_description):
        """
        Function to verify the alert after create
        :param message_description:
        """
        self.open_message_tab()
        if message_description == self.webDriver.get_text(element=self.locators.get_message_description,
                                                          locator_type='xpath'):
            self.webDriver.allure_attach_jpeg('added_message')
            logger.debug("Message Description is matching after creation ")
        else:
            self.webDriver.allure_attach_jpeg('error_added_message')
            raise Exception("Message Description is not matching after creation ")

    def add_additional_visit(self, additional_visit_reason, ship_visit_time, reviewer, side):
        """
        Click on add visit on Visitor Details page
        :param additional_visit_reason:
        :param ship_visit_time:
        :param reviewer:
        """
        self.webDriver.click(element=self.locators.add_visit, locator_type='xpath')
        self.webDriver.wait_for(5)
        new_date = datetime.strptime(ship_visit_time, "%m/%d/%Y")
        ship_visit_time = str(new_date + timedelta(days=1))
        ship_visit_time = datetime.strptime(ship_visit_time, "%Y-%m-%d %H:%M:%S").strftime("%m/%d/%Y")
        self.webDriver.set_text(element=self.locators.visit_date, locator_type='xpath', text=ship_visit_time)
        self.webDriver.wait_for(5)
        self.webDriver.set_text(element=self.locators.visit_purpose, locator_type='xpath', text=additional_visit_reason)
        self.webDriver.enter_data_in_textbox_using_action(element=self.locators.boarding_type,
                                                          locator_type='xpath',
                                                          text='General Boarding')
        self.webDriver.wait_for(5)
        if side == 'shore':
            self.webDriver.enter_data_in_textbox_using_action(element=self.locators.reviewer, locator_type='xpath', text=reviewer)
        self.webDriver.allure_attach_jpeg('add_new_visit')
        self.webDriver.click(element=self.locators.submit, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.open_visit_tab()
        self.webDriver.allure_attach_jpeg('visit_tab_after_add_visit')
        self.webDriver.click(element=self.locators.click_visit_arrow, locator_type='xpath')
        visit_reason = self.webDriver.get_text(element=self.locators.visit_purpose_value,
                                               locator_type='xpath')
        if visit_reason == additional_visit_reason:
            logger.debug("New visit is added")
            self.webDriver.allure_attach_jpeg('new_visit_added')
        else:
            self.webDriver.allure_attach_jpeg('error_in_new_visit')
            raise Exception("New visit is not added")
        self.webDriver.click(element=self.locators.cancel, locator_type='xpath')

    def verify_visitor_export(self, file_name):
        """
        Function to click on export functionality
        :param file_name:
        """
        self.webDriver.allure_attach_jpeg('before_export_click')
        self.webDriver.scroll_complete_page_top()
        self.webDriver.wait_for(1)
        self.webDriver.click(element=self.locators.export_icon, locator_type='xpath')
        self.webDriver.wait_for(1)
        self.webDriver.click(element=self.locators.export_excel, locator_type='xpath')
        self.webDriver.wait_for(1)
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('after_export_click')

        if os.environ.get('EXECUTOR_NUMBER', None):
            flag = self.webDriver.grid_download_verification(file_name=file_name)
        else:
            download_path = os.getcwd() + '/downloaded_files/'
            if not os.path.isdir(download_path):
                os.makedirs(download_path)
            file_path = pathlib.Path(download_path + f"{file_name}.xls")
            flag = self.webDriver.check_local_system_download(file_path)

        if flag:
            logger.debug("file is downloaded successfully")
        else:
            raise Exception("File is not downloaded successfully")

    def verify_bulk_approve(self, file_name, names):
        """
        Function to Verify the bulk approval status in excel
        :param file_name:
        :param names:
        """
        self.webDriver.get_file_content()
        download_path = f"{os.getcwd()}/{file_name}"
        if not os.path.isdir(download_path):
            os.makedirs(download_path)
        if len(download_path) > 0:
            wb = openpyxl.Workbook()
            ws = wb.active

            with open(f"{download_path}.csv") as f:
                reader = csv.reader(f, delimiter=':')
                for row in reader:
                    ws.append(row)

            wb.save(f"{download_path}.xlsx")
            wb_obj = openpyxl.load_workbook(f"{download_path}.xlsx")
            sheet_obj = wb_obj.active
            rows_count = sheet_obj.max_row
            for i in range(2, rows_count):
                details = sheet_obj.cell(i, 1).value.split(',')
                for _name in names:
                    if _name in details:
                        assert details[
                                   12].upper() == "APPROVED", f'Bulk Approved visitor {_name} is not approved in exported excel !!'
        else:
            raise Exception("File is not transferred successfully")

        os.remove(f"{download_path}.xlsx")

    # def import_visitors(self, first_name, last_name, country, doc_id, visit_date):

    def verify_bulk_visitor_approved(self, test_data):
        """
        Function to verify bulk visitors approved
        :param test_data:
        :return:
        """
        if self.webDriver.is_element_display_on_screen(element=self.locators.no_records, locator_type='xpath'):
            pytest.skip("No records found for visitors")
        else:
            visitors = self.webDriver.get_elements(element=self.locators.visitors, locator_type='xpath')
            for i in range(1, len(visitors)):
                id = self.webDriver.get_text(element=self.locators.column_value_in_row % (i,test_data['ID Number_position']), locator_type='xpath')
                if id in test_data['visitors_id_number']:
                    self.webDriver.get_text(
                        element=self.locators.column_value_in_row % (i, test_data['Visit Status_position'])) == "Approved"
                else:
                    raise Exception("Bulk visitor approve failed")

    def get_doc_type_name(self):
        """
        Function to get the doc type name
        """
        return self.webDriver.get_text(element=self.locators.details_ID_type, locator_type='xpath')

    def check_availability_of_visitor_details_screen(self):
        """
        Function to get the doc type name
        """
        self.webDriver.wait_till_element_appear_on_screen(element=self.locators.visitor_details,
                                                          locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')

    def blank_list(self):
        """
        Verification of blank list
        """
        if self.webDriver.is_element_display_on_screen(element=self.locators.no_records, locator_type='xpath'):
            return True
        else:
            return False

    def check_new_visitor_button_is_disabled(self):
        """
        Function to check the new visitor button is disabled
        """
        if self.webDriver.is_element_display_on_screen(element=self.locators.new_visitor_disabled_button,
                                                       locator_type='xpath'):
            return True
        else:
            return False

    def get_current_portcode(self, test_data):
        """
        Function to get current portcode
        :return:
        """
        day = date.today().strftime("%d")
        current_day_index = test_data['active_voyage_date_list'].index(day)
        test_data['current_portcode'] = test_data['port_list'][current_day_index]
