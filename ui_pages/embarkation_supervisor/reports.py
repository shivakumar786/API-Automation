__author__ = 'prahlad.sharma'

import csv

import openpyxl

from virgin_utils import *
import pathlib


class Reports(General):
    """
    Page Class for reports in Embarkation Supervisor
    """

    def __init__(self, web_driver):
        """
        To Initialize the locators of reports page
        :param web_driver:
        """
        super().__init__()
        self.webDriver = web_driver
        self.locators = self.dict_to_ns({
            "loader": "//div[@class='LoaderContainer__div']//img",
            "filter_start_date":"//div[@class='column is-5'][1]/div/div/div[@class='react-datepicker__input-container']/input",
            "filter_end_date": "//div[@class='column is-5'][2]/div/div/div[@class='react-datepicker__input-container']/input",
            "filter_date_blank":"//input[@class='react-datepicker-ignore-onclickoutside']",
            "report_header_title": "//div[@class='header']//span[contains(text(),'Reports')]",
            "exception_report": "//span[contains(text(),'Exception Report')]",
            "location_report": "//span[contains(text(),'Location History Report')]",
            "movement_report": "//span[contains(text(),'Movement Report')]",
            "cbp_report": "//span[contains(text(),'CBP Report')]",
            "ipm_report": "//span[contains(text(),'IPM Report')]",
            "missing_picture_tab": "//button[contains(text(),'Missing Picture')]",
            "expire_card": "//button[contains(text(),'Expired Card')]",
            "Not_Checked_in_Onboard": "//button[contains(text(),'Not Checked-in, Onboard')]",
            "Checked_in_Not_Onboard": "//button[contains(text(),'Checked-in, Not Onboard')]",
            "Visitor_Onboard": "//button[contains(text(),'Visitor Onboard')]",
            "Wearable_Assigned": "//button[contains(text(),'Wearable Assigned')]",
            "export_icon": "//div[@id='DownloadContainer__div']//img[contains(@class,'SVGContainer')]",
            "export_csv": "//span[contains(text(),'Export CSV')]",
            "sailor_not_cleared": "//span[contains(text(),'Sailors')]/../following-sibling::div/div/span[starts-with(text(),'Not Cleared')]",
            "sailor_cleared": "//span[contains(text(),'Sailors')]/../following-sibling::div/div/span[starts-with(text(),'Cleared')]",
            "sailor_total": "//span[contains(text(),'Sailors')]/../following-sibling::div/div/span[starts-with(text(),'Sailor Total')]",
            "sailor_total_count": "//span[contains(text(),'Sailors')]/../following-sibling::div/div/span[starts-with(text(),'Sailor Total')]/../preceding-sibling::div",
            "crew_not_cleared": "//span[contains(text(),'Crew')]/../following-sibling::div/div/span[starts-with(text(),'Not Cleared')]",
            "crew_cleared": "//span[contains(text(),'Crew')]/../following-sibling::div/div/span[starts-with(text(),'Cleared')]",
            "crew_total": "//span[contains(text(),'Crew')]/../following-sibling::div/div/span[starts-with(text(),'Crew Total')]",
            "crew_total_count": "//span[contains(text(),'Crew')]/../following-sibling::div/div/span[starts-with(text(),'Crew Total')]/../preceding-sibling::div",
            "filter": "//div[@class='filters']//img[contains(@class,'SVGContainer')]",
            "boarding_status": "//button[contains(text(),'Boarding Status')]",
            "onboard": "//label[contains(text(),'Onboard')]",
            "ashore": "//label[contains(text(),'Ashore')]",
            "clearance_status": "//button[contains(text(),'Clearance Status')]",
            "alert_clear": "//label[contains(text(),'Alert Cleared')]",
            "alert_not_clear": "//label[contains(text(),'Alert Not Cleared')]",
            "department": "//button[contains(text(),'Department')]",
            "select_department": "//span[contains(text(),'%s')]",
            "apply": "//span[contains(text(),'Apply')]",
            "cancel": "//span[contains(text(),'Cancel')]",
            "clear_all": "//span[contains(text(),'Clear All')]",
            "no_records": "//span[contains(text(),'No Record(s) Available.')]",
            "header_name": "//thead/tr/th",
            "column_value_in_row": "//tbody/tr[%s]/td[%d]",
            "report_header": "//span[contains(text(),'%s')]/..",
            "assignment": "//button[contains(text(),'Assignment')]",
            "without_assignment": "//label[contains(text(),'Without Assignment')]",
            "report_count": "//div[@class='count']",
            "gender_filter": "//button[contains(text(),'Gender')]",
            "select_gender_female": "//label[text()='Female']",

        })

    def verification_of_report_page(self):
        """
        To check the availability of reports page
        """
        self.webDriver.explicit_visibility_of_element(element=self.locators.report_header_title, locator_type='xpath',
                                                      time_out=60)
        screen_title = self.webDriver.get_text(element=self.locators.report_header_title, locator_type='xpath')
        if screen_title == "Reports":
            logger.debug("User is land on Reports page")
        else:
            raise Exception("user is not on Reports page")

    def open_exception_report(self):
        """
        Function to open the exception reports
        """
        self.webDriver.click(element=self.locators.exception_report, locator_type='xpath')
        self.webDriver.explicit_invisibility_of_element(element=self.locators.loader, locator_type='xpath',
                                                        time_out=60)
        self.webDriver.allure_attach_jpeg('exception_report_landing')

    def open_location_history_report(self):
        """
        Function to open the location reports
        """
        self.webDriver.click(element=self.locators.location_report, locator_type='xpath')
        self.webDriver.explicit_invisibility_of_element(element=self.locators.loader, locator_type='xpath',
                                                        time_out=60)
        self.webDriver.allure_attach_jpeg('location_report_landing')

    def open_movement_report(self):
        """
        Function to open the movement reports
        """
        self.webDriver.click(element=self.locators.movement_report, locator_type='xpath')
        self.webDriver.explicit_invisibility_of_element(element=self.locators.loader, locator_type='xpath',
                                                        time_out=60)
        self.webDriver.allure_attach_jpeg('movement_report_landing')

    def open_cbp_report(self):
        """
        Function to open the cbp reports
        """
        self.webDriver.click(element=self.locators.cbp_report, locator_type='xpath')
        self.webDriver.explicit_invisibility_of_element(element=self.locators.loader, locator_type='xpath',
                                                        time_out=60)
        self.webDriver.allure_attach_jpeg('cbp_report_landing')

    def open_ipm_report(self):
        """
        Function to open the IPM reports
        """
        self.webDriver.click(element=self.locators.ipm_report, locator_type='xpath')
        self.webDriver.explicit_invisibility_of_element(element=self.locators.loader, locator_type='xpath',
                                                        time_out=60)
        self.webDriver.allure_attach_jpeg('ipm_report_landing')

    def get_missing_picture_count(self):
        """
        Function to get the missing picture tab count and click on tab
        """
        self.webDriver.click(element=self.locators.missing_picture_tab, locator_type='xpath')
        missing_count = self.webDriver.get_text(element=self.locators.missing_picture_tab, locator_type='xpath')
        count = missing_count.rsplit(' ', 1)[1].replace('(', '').replace(')', '')
        return int(count)

    def get_expired_count(self):
        """
        Function to get the expired tab count and click on tab
        """
        self.webDriver.click(element=self.locators.expire_card, locator_type='xpath')
        missing_count = self.webDriver.get_text(element=self.locators.expire_card, locator_type='xpath')
        count = missing_count.rsplit(' ', 1)[1].replace('(', '').replace(')', '')
        return int(count)

    def get_not_checkedin_onboard_count(self):
        """
        Function to get the Not Checked in On board tab count and click on tab
        """
        self.webDriver.click(element=self.locators.Not_Checked_in_Onboard, locator_type='xpath')
        missing_count = self.webDriver.get_text(element=self.locators.Not_Checked_in_Onboard, locator_type='xpath')
        count = missing_count.rsplit(' ', 1)[1].replace('(', '').replace(')', '')
        return int(count)

    def get_checked_in_not_onboard_count(self):
        """
        Function to get the Checked in not On board tab count and click on tab
        """
        self.webDriver.click(element=self.locators.Checked_in_Not_Onboard, locator_type='xpath')
        missing_count = self.webDriver.get_text(element=self.locators.Checked_in_Not_Onboard, locator_type='xpath')
        count = missing_count.rsplit(' ', 1)[1].replace('(', '').replace(')', '')
        return int(count)

    def get_visitor_on_board_count(self):
        """
        Function to get the visitor On board tab count and click on tab
        """
        self.webDriver.click(element=self.locators.Visitor_Onboard, locator_type='xpath')
        missing_count = self.webDriver.get_text(element=self.locators.Visitor_Onboard, locator_type='xpath')
        count = missing_count.rsplit(' ', 1)[1].replace('(', '').replace(')', '')
        return int(count)

    def get_wearable_assigned_count(self):
        """
        Function to get the wearable assigned tab count and click on tab
        """
        self.webDriver.click(element=self.locators.Wearable_Assigned, locator_type='xpath')
        missing_count = self.webDriver.get_text(element=self.locators.Wearable_Assigned, locator_type='xpath')
        count = missing_count.rsplit(' ', 1)[1].replace('(', '').replace(')', '')
        return int(count)

    def export_and_verify_reports(self, file_name):
        """
        Function to export the reports
        :param file_name:
        """
        rows_count = None
        worksheet = None
        workbook = None
        self.webDriver.click(element=self.locators.export_icon, locator_type='xpath')
        self.webDriver.wait_for(1)
        self.webDriver.click(element=self.locators.export_csv, locator_type='xpath')
        self.webDriver.wait_for(30)
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('after_export_click')

        self.webDriver.get_file_content()
        download_path = f"{os.getcwd()}/downloaded_files/{file_name}"
        if not os.path.isdir(download_path):
            os.makedirs(download_path)
        if len(download_path) > 0:
            wb = openpyxl.Workbook()
            ws = wb.active

            with open(f"{download_path}.csv") as f:
                reader = csv.reader(f, delimiter=':')
                for row in reader:
                    ws.append(row)

            wb.save(f"{download_path}.xlsx")
            wb_obj = openpyxl.load_workbook(f"{download_path}.xlsx")
            sheet_obj = wb_obj.active
            rows_count = sheet_obj.max_row - 1
            os.remove(f"{download_path}.xlsx")
            os.remove(f"{download_path}.csv")
        else:
            raise Exception("File is not transferred successfully")

        return rows_count

    def check_availability_of_fields_in_cbp(self):
        """
        Function to check the availability of fields in CBP reports
        """
        if self.webDriver.is_element_display_on_screen(element=self.locators.sailor_not_cleared, locator_type='xpath'):
            logger.debug("Sailor Not Cleared available on screen")
        else:
            raise Exception("Sailor Not Cleared not available on screen")

        if self.webDriver.is_element_display_on_screen(element=self.locators.sailor_total, locator_type='xpath'):
            logger.debug("Sailor total available on screen")
        else:
            raise Exception("Sailor total not available on screen")

        if self.webDriver.is_element_display_on_screen(element=self.locators.crew_not_cleared, locator_type='xpath'):
            logger.debug("Crew Not Cleared available on screen")
        else:
            raise Exception("Sailor Not Cleared not available on screen")

        if self.webDriver.is_element_display_on_screen(element=self.locators.crew_total, locator_type='xpath'):
            logger.debug("Sailor total available on screen")
        else:
            raise Exception("Sailor total not available on screen")

    def get_total_sailor_count_cbp(self):
        """
        Function to get the sailor count
        """
        return self.webDriver.get_text(element=self.locators.sailor_total_count, locator_type='xpath')

    def get_total_crew_count(self):
        """
        Function to get the crew count
        """
        return self.webDriver.get_text(element=self.locators.crew_total_count, locator_type='xpath')

    def click_filter_option(self):
        """
        Function to click on filter option
        """
        self.webDriver.click(element=self.locators.filter, locator_type='xpath')
        self.webDriver.wait_for(2)

    def select_boarding_status(self):
        """
        Function to select the boarding status
        """
        self.webDriver.click(element=self.locators.boarding_status, locator_type='xpath')
        self.webDriver.click(element=self.locators.onboard, locator_type='xpath')

    def select_clearance_status(self):
        """
        Function to select the boarding status
        """
        self.webDriver.click(element=self.locators.clearance_status, locator_type='xpath')
        self.webDriver.click(element=self.locators.alert_not_clear, locator_type='xpath')

    def select_gender(self):
        """
        Function to select the boarding status
        """
        self.webDriver.click(element=self.locators.gender_filter, locator_type='xpath')
        self.webDriver.click(element=self.locators.select_gender_female, locator_type='xpath')


    def select_department(self):
        """
        Function to select the department
        """
        self.webDriver.click(element=self.locators.department, locator_type='xpath')
        self.webDriver.click(element=self.locators.select_department % 'Casino', locator_type='xpath')
        self.webDriver.click(element=self.locators.select_department % 'Galley', locator_type='xpath')
        self.webDriver.click(element=self.locators.select_department % 'Gift Shop', locator_type='xpath')

    def click_apply(self):
        """
        Function to click on Apply button
        """
        self.webDriver.click(element=self.locators.apply, locator_type='xpath')
        self.webDriver.wait_till_element_disappear_from_screen(element=self.locators.loader, locator_type='xpath')
        self.webDriver.allure_attach_jpeg('cbp_report_after_apply')

    def blank_list(self):
        """
        Verification of blank list
        """
        if self.webDriver.is_element_display_on_screen(element=self.locators.no_records, locator_type='xpath'):
            return True
        else:
            return False

    def set_column_name(self, test_data):
        """
        Function to get the column name
        :param test_data:
        """
        column_number = 0
        total_column = self.webDriver.get_elements(element=self.locators.header_name, locator_type='xpath')
        for column in total_column:
            column_number = column_number + 1
            test_data[f"{column.text}_position"] = column_number

    def check_movement_availability(self, test_data):
        """
        Function to check the availability of movement
        :param test_data:
        """
        movement_type = self.webDriver.get_text(self.locators.column_value_in_row % (1, test_data['Movement_position']),
                                                locator_type='xpath')
        if movement_type == 'ONBOARD' or movement_type == 'ASHORE':
            logger.debug("Correct status display")
        else:
            raise Exception("Correct status is not display")

    def get_total_location_count(self):
        """
        Function to get the sailor count
        :param name:
        """
        value = self.webDriver.get_text(element=self.locators.report_count, locator_type='xpath')
        location_count = value.replace(')', '').replace('(', '')
        return int(location_count)

    def get_total_report_count(self, name):
        """
        Function to get the sailor count
        :param name:
        """
        value = self.webDriver.get_text(element=self.locators.report_header % name, locator_type='xpath')
        location_count = value.rsplit(' ', 1)[1].replace('(', '').replace(')', '')
        return int(location_count)

    def select_assignment_status(self):
        """
        Function to select the assignment status
        """
        self.webDriver.click(element=self.locators.assignment, locator_type='xpath')
        self.webDriver.click(element=self.locators.without_assignment, locator_type='xpath')

